import openpyxl
import Levenshtein

def find_closest_match(target, options):
    closest_match = None
    min_distance = float('inf')
    for option in options:
        distance = Levenshtein.distance(target, option)
        if distance < min_distance:
            min_distance = distance
            closest_match = option
    return closest_match

def update_excel(workbook_path, sheet_name, student_id, new_score):
    # Load the workbook
    workbook = openpyxl.load_workbook(workbook_path)

    # Find the closest match for the sheet name
    closest_sheet_name = find_closest_match(sheet_name, workbook.sheetnames)

    if closest_sheet_name is None:
        print(f"No close matches found for sheet name '{sheet_name}'.")
        workbook.close()
        return


    # Select the closest matching worksheet
    sheet = workbook[closest_sheet_name]

    # Collect all student IDs in column C and ensure they are strings
    student_ids = []

    for row in sheet.iter_rows(min_row=1, min_col=3, max_col=3, values_only=True):
        if row[0] is not None:  # Only add non-empty cells
            student_ids.append(str(row[0]))

    # Find the closest match for the student ID
    closest_id = find_closest_match(student_id, student_ids)

    if closest_id is None:
        print(f"No close matches found for student ID '{student_id}'.")
    else:
        id_found = False
        for row in sheet.iter_rows(min_row=1, min_col=3, max_col=6):
            cell = row[0]  # Cell in column C
            if str(cell.value) == closest_id:
                id_row_index = cell.row
                sheet[f'F{id_row_index}'] = new_score  # Update the score in column F
                id_found = True
                break

        if id_found:
            # Save the workbook
            workbook.save(workbook_path)
            print(f"Score of student ID '{closest_id}' updated to '{new_score}' in sheet '{closest_sheet_name}'.")
        else:
            print(f"Unexpected error: Student ID '{closest_id}' not found after searching.")

    # Close the workbook
    workbook.close()
